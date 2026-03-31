# -*- coding: utf-8 -*-
"""
2025 年华南地区新能源发电量数据生成与可视化
华南地区：广东、广西、海南

v2.0 - 添加配置化和数据验证功能
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

# 导入配置
from config import DataConfig, ChartConfig, DEFAULT_DATA_CONFIG, DEFAULT_CHART_CONFIG

# 设置中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


def validate_data(data_config: DataConfig) -> bool:
    """验证数据完整性"""
    validation = data_config.validate()
    all_passed = all(validation.values())
    
    print("\n=== 数据验证结果 ===")
    for check, passed in validation.items():
        status = "[OK]" if passed else "[FAIL]"
        print(f"  {status} {check}")
    
    if not all_passed:
        raise ValueError("数据验证失败，请检查配置")
    
    print("\n[OK] 数据验证通过")
    return True


def generate_excel(data_config: DataConfig, output_path: str) -> None:
    """生成 Excel 数据文件"""
    months = ['1 月', '2 月', '3 月', '4 月', '5 月', '6 月', 
              '7 月', '8 月', '9 月', '10 月', '11 月', '12 月']
    
    total_data = data_config.get_total()
    
    df = pd.DataFrame({
        '月份': months,
        '广东': data_config.guangdong_data,
        '广西': data_config.guangxi_data,
        '海南': data_config.hainan_data,
        '华南总计': total_data
    })
    
    print("\n=== 2025 年华南地区新能源月度发电量数据 ===")
    print(df.to_string(index=False))
    
    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "月度发电量数据"
    
    # 样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = ['月份', '广东 (亿 kWh)', '广西 (亿 kWh)', '海南 (亿 kWh)', '华南总计 (亿 kWh)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 数据
    data_alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            if col_idx == 5:
                cell.font = Font(bold=True)
    
    # 列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    # 汇总行
    annual = data_config.get_annual_total()
    total_row = 14
    ws.cell(row=total_row, column=1, value="全年合计").font = Font(bold=True)
    for col_idx, (province, val) in enumerate(annual.items(), 2):
        cell = ws.cell(row=total_row, column=col_idx, value=f"{val:.1f}")
        cell.font = Font(bold=True)
        cell.border = thin_border
    
    wb.save(output_path)
    print(f"\n[OK] Excel 文件已保存：{output_path}")


def generate_charts(data_config: DataConfig, chart_config: ChartConfig, fig_dir: str) -> None:
    """生成可视化图表"""
    months = ['1 月', '2 月', '3 月', '4 月', '5 月', '6 月', 
              '7 月', '8 月', '9 月', '10 月', '11 月', '12 月']
    
    total_data = data_config.get_total()
    
    df = pd.DataFrame({
        '月份': months,
        '广东': data_config.guangdong_data,
        '广西': data_config.guangxi_data,
        '海南': data_config.hainan_data,
        '华南总计': total_data
    })
    
    # 图 1: 折线图
    fig1, ax1 = plt.subplots(figsize=chart_config.line_chart_size)
    
    for province in ['广东', '广西', '海南', '华南总计']:
        ax1.plot(months, df[province], marker='o', linewidth=2.5, 
                 markersize=8, label=province, color=chart_config.colors.get(province, '#000000'))
    
    ax1.set_xlabel('月份', fontsize=12, fontweight='bold')
    ax1.set_ylabel('发电量 (亿千瓦时)', fontsize=12, fontweight='bold')
    ax1.set_title('2025 年华南地区新能源月度发电量趋势', fontsize=14, fontweight='bold', pad=15)
    ax1.legend(fontsize=11, loc='upper right')
    ax1.grid(True, alpha=0.3, linestyle='--')
    plt.xticks(rotation=0)
    
    # 数据标签
    for province in ['广东', '广西', '海南']:
        for i, val in enumerate(df[province]):
            ax1.annotate(f'{val:.1f}', (i, val), textcoords="offset points", 
                        xytext=(0,5), ha='center', fontsize=7)
    
    plt.tight_layout()
    line_chart_path = f"{fig_dir}/月度发电量折线图.png"
    plt.savefig(line_chart_path, dpi=chart_config.dpi, bbox_inches='tight')
    plt.close()
    print(f"[OK] 折线图已保存：{line_chart_path}")
    
    # 图 2: 饼图
    fig2, ax2 = plt.subplots(figsize=chart_config.pie_chart_size)
    annual = data_config.get_annual_total()
    province_totals = [annual['广东'], annual['广西'], annual['海南']]
    province_names = ['广东', '广西', '海南']
    pie_colors = [chart_config.colors['广东'], chart_config.colors['广西'], chart_config.colors['海南']]
    
    explode = (0.05, 0, 0)
    wedges, texts, autotexts = ax2.pie(province_totals, explode=explode, 
                                         labels=province_names, 
                                         autopct='%1.1f%%',
                                         colors=pie_colors,
                                         shadow=True, 
                                         startangle=90,
                                         textprops={'fontsize': 12, 'weight': 'bold'})
    
    for autotext in autotexts:
        autotext.set_color('white')
        
    ax2.set_title('2025 年华南三省新能源发电量占比', fontsize=14, fontweight='bold', pad=20)
    
    total_sum = sum(province_totals)
    legend_text = f'\n全年总发电量：{total_sum:.1f} 亿千瓦时'
    fig2.text(0.5, 0.02, legend_text, ha='center', fontsize=11, 
              bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.5))
    
    plt.tight_layout()
    pie_chart_path = f"{fig_dir}/全年发电量占比饼图.png"
    plt.savefig(pie_chart_path, dpi=chart_config.dpi, bbox_inches='tight')
    plt.close()
    print(f"[OK] 饼图已保存：{pie_chart_path}")
    
    # 图 3: 柱状图
    fig3, ax3 = plt.subplots(figsize=chart_config.bar_chart_size)
    x = np.arange(len(months))
    width = 0.25
    
    bars1 = ax3.bar(x - width, data_config.guangdong_data, width, label='广东', color=chart_config.colors['广东'])
    bars2 = ax3.bar(x, data_config.guangxi_data, width, label='广西', color=chart_config.colors['广西'])
    bars3 = ax3.bar(x + width, data_config.hainan_data, width, label='海南', color=chart_config.colors['海南'])
    
    ax3.set_xlabel('月份', fontsize=12, fontweight='bold')
    ax3.set_ylabel('发电量 (亿千瓦时)', fontsize=12, fontweight='bold')
    ax3.set_title('2025 年华南三省新能源月度发电量对比', fontsize=14, fontweight='bold', pad=15)
    ax3.set_xticks(x)
    ax3.set_xticklabels(months)
    ax3.legend(fontsize=11)
    ax3.grid(True, alpha=0.3, axis='y', linestyle='--')
    
    for bars in [bars1, bars2, bars3]:
        for bar in bars:
            height = bar.get_height()
            ax3.annotate(f'{height:.1f}',
                        xy=(bar.get_x() + bar.get_width() / 2, height),
                        xytext=(0, 3),
                        textcoords="offset points",
                        ha='center', va='bottom', fontsize=7)
    
    plt.tight_layout()
    bar_chart_path = f"{fig_dir}/各省月度对比柱状图.png"
    plt.savefig(bar_chart_path, dpi=chart_config.dpi, bbox_inches='tight')
    plt.close()
    print(f"[OK] 柱状图已保存：{bar_chart_path}")


def main():
    """主函数"""
    print("=" * 60)
    print("2025 年华南地区新能源发电量数据生成与可视化 v2.0")
    print("=" * 60)
    
    # 加载配置
    data_config = DEFAULT_DATA_CONFIG
    chart_config = DEFAULT_CHART_CONFIG
    
    # 数据验证
    validate_data(data_config)
    
    # 路径
    base_dir = r"D:\AI-agent\openclaw-apri\projects\2025-新能源华南双碳报告"
    excel_path = f"{base_dir}\\data\\2025 年华南新能源发电量.xlsx"
    fig_dir = f"{base_dir}\\charts"
    
    # 生成 Excel
    generate_excel(data_config, excel_path)
    
    # 生成图表
    generate_charts(data_config, chart_config, fig_dir)
    
    # 输出年度统计
    annual = data_config.get_annual_total()
    print("\n=== 年度统计 ===")
    for province, total in annual.items():
        print(f"  {province}: {total:.1f} 亿千瓦时")
    
    print("\n=== 数据生成与可视化完成 ===")


if __name__ == '__main__':
    main()